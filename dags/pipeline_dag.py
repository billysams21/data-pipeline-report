from __future__ import annotations
import datetime
import pendulum
import pandas as pd
from zipfile import ZipFile
from io import BytesIO, TextIOWrapper
from sqlalchemy import create_engine
import os

from airflow.models.dag import DAG
from airflow.operators.python import PythonOperator
from airflow.providers.postgres.hooks.postgres import PostgresHook
from airflow.sensors.python import PythonSensor

from cosmos import DbtTaskGroup, ProjectConfig, ProfileConfig
from cosmos.profiles import PostgresUserPasswordProfileMapping

from openpyxl.styles import Font, PatternFill
from function.pdf_generator import generate_pdf_report
from function.model_trainer import train_and_evaluate_models

def run_ml_pipeline_with_report(postgres_conn_id: str):
    """
    Runs the complete ML pipeline: train models and generate PDF report with results.
    """
    try:
        # Step 1: Train and evaluate models
        print("Starting machine learning model training...")
        model_chart_path = train_and_evaluate_models(postgres_conn_id)
        print(f"Model training completed. Chart saved to: {model_chart_path}")
        
        # Step 2: Generate PDF report with ML results
        print("Generating PDF report with ML results...")
        pdf_report_path = generate_pdf_report(postgres_conn_id, model_chart_path)
        print(f"PDF report generated: {pdf_report_path}")
        
        return model_chart_path, pdf_report_path
        
    except Exception as e:
        print(f"Error in ML pipeline: {e}")
        try:
            pdf_report_path = generate_pdf_report(postgres_conn_id)
            return None, pdf_report_path
        except Exception as pdf_error:
            print(f"Error generating PDF report: {pdf_error}")
            raise

# Path
BASE_PATH = "/opt/airflow/dags/data"
OUTPUT_PATH = "/opt/airflow/output"
MAIN_ZIP_FILE = f"{BASE_PATH}/bank.zip"
ADDITIONAL_ZIP_FILE = f"{BASE_PATH}/bank-additional.zip"
DBT_PROJECT_PATH = f"{os.environ['AIRFLOW_HOME']}/dbt"

# Cosmos
profile_config = ProfileConfig(
    profile_name="analytics_project",
    target_name="dev",
    profile_mapping=PostgresUserPasswordProfileMapping(
        conn_id="postgres_default",
        profile_args={"schema": "public"},
    )
)

# Operator
def check_file_exists(file_path: str):
    exists = os.path.exists(file_path)
    print(f"Checking file existence: {file_path} - {'EXISTS' if exists else 'NOT FOUND'}")
    if not exists:
        # List directory contents for debugging
        directory = os.path.dirname(file_path)
        if os.path.exists(directory):
            files = os.listdir(directory)
            print(f"Directory {directory} contents: {files}")
        else:
            print(f"Directory {directory} does not exist")
    return exists

def extract_and_load_csv(zip_path: str, csv_filename: str, table_name: str, postgres_conn_id: str):
    """
    Fungsi untuk ekstrak CSV dari ZIP dan ke PostgreSQL
    """
    print(f"Ekstrak {csv_filename} dari {zip_path}")
    
    try:
        # Koneksi ke PostgreSQL dengan Hook
        hook = PostgresHook(postgres_conn_id=postgres_conn_id)
        
        with ZipFile(zip_path, 'r') as zip_ref:
            zip_contents = zip_ref.namelist()
            # print(f"Isi ZIP file: {zip_contents}")
            
            # Cari file CSV dalam ZIP
            csv_file_path = None
            for file_path in zip_contents:
                if file_path.endswith(csv_filename) or file_path.endswith(f"/{csv_filename}"):
                    csv_file_path = file_path
                    break
            
            # Jika tidak ditemukan dengan nama exact, cari dengan pattern
            if csv_file_path is None:
                for file_path in zip_contents:
                    if csv_filename in file_path and file_path.endswith('.csv'):
                        csv_file_path = file_path
                        break
            
            if csv_file_path is None:
                raise FileNotFoundError(f"File {csv_filename} tidak ditemukan dalam ZIP. Isi ZIP: {zip_contents}")
            
            print(f"Menggunakan file: {csv_file_path}")
            
            with zip_ref.open(csv_file_path, 'r') as csv_file:
                df = pd.read_csv(TextIOWrapper(csv_file, 'utf-8'), sep=';')
                print(f"Berhasil membaca {len(df)} baris dari {csv_file_path}")  
                
                # Bersihkan nama kolom
                df.columns = df.columns.str.replace('.', '_', regex=False)
                df.columns = df.columns.str.replace('"', '', regex=False)
                df.columns = df.columns.str.strip()
                
                hook.run("CREATE SCHEMA IF NOT EXISTS staging;")
                
                columns_def = []
                for col in df.columns:
                    columns_def.append(f'"{col}" TEXT')
                
                create_table_sql = f"""
                DROP TABLE IF EXISTS staging.{table_name} CASCADE;
                CREATE TABLE staging.{table_name} (
                    {', '.join(columns_def)}
                );
                """
                hook.run(create_table_sql)
                
                # Insert data dalam batch
                batch_size = 10000
                total_rows = len(df)
                
                for start_idx in range(0, total_rows, batch_size):
                    end_idx = min(start_idx + batch_size, total_rows)
                    batch_df = df.iloc[start_idx:end_idx]
                    
                    values_list = []
                    for _, row in batch_df.iterrows():
                        values = []
                        for value in row:
                            if pd.isna(value):
                                values.append('NULL')
                            else:
                                # Escape single quotes
                                escaped_value = str(value).replace("'", "''")
                                values.append(f"'{escaped_value}'")
                        values_list.append(f"({', '.join(values)})")
                    
                    if values_list:
                        insert_sql = f"""
                        INSERT INTO staging.{table_name} ({', '.join([f'"{col}"' for col in df.columns])})
                        VALUES {', '.join(values_list)};
                        """
                        hook.run(insert_sql)
                    
                    print(f"Batch {start_idx//batch_size + 1}: Memuat baris {start_idx+1}-{end_idx}")
                
                print(f"Berhasil memuat {len(df)} baris ke tabel staging.{table_name}")
                
    except Exception as e:
        print(f"Error dalam extract_and_load_csv: {str(e)}")
        raise

def generate_excel_report(postgres_conn_id: str):
    """
    Mengambil data dari dbt dan membuat Excel.
    """
    os.makedirs(OUTPUT_PATH, exist_ok=True)
    hook = PostgresHook(postgres_conn_id=postgres_conn_id)
    
    # Mengambil data dari dbt models
    df_data = hook.get_pandas_df(sql="SELECT * FROM public.report_full_dump")
    df_dict = hook.get_pandas_df(sql="SELECT * FROM public.dim_variables")
    
    output_file = f"{OUTPUT_PATH}/report.xlsx"
    
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df_data.to_excel(writer, sheet_name='Data', index=False)
        df_dict.to_excel(writer, sheet_name='Data Dictionary', index=False)
        
        workbook = writer.book
        # Sheet 'Data'
        sheet_data = writer.sheets['Data']
        sheet_data.freeze_panes = 'A2' # Freeze
        for col_idx, col in enumerate(sheet_data.columns):
            col_name = df_data.columns[col_idx]
            max_length = max(df_data[col_name].astype(str).map(len).max(), len(str(col[0].value)))
            sheet_data.column_dimensions[col[0].column_letter].width = max_length + 2
        for cell in sheet_data[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDEEFF", end_color="DDEEFF", fill_type="solid")

        # Sheet 'Data Dictionary'
        sheet_dict = writer.sheets['Data Dictionary']
        sheet_dict.freeze_panes = 'A2' # Freeze
        for col_idx, col in enumerate(sheet_dict.columns):
            col_name = df_dict.columns[col_idx]
            max_length = max(df_dict[col_name].astype(str).map(len).max(), len(str(col[0].value)))
            sheet_dict.column_dimensions[col[0].column_letter].width = max_length + 2
        for cell in sheet_dict[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDEEFF", end_color="DDEEFF", fill_type="solid")
    print(f"Laporan Excel berhasil dibuat di: {output_file}")

# DAG
with DAG(
    dag_id="data_ingestion_pipeline",
    start_date=pendulum.datetime(2025, 1, 1, tz="UTC"),
    schedule="*/55 * * * *",
    catchup=False,
    tags=["ingestion", "dbt", "reporting"],
) as dag:

    # Task Sensor
    wait_for_main_file = PythonSensor(
        task_id="wait_for_main_file",
        python_callable=check_file_exists,
        op_kwargs={"file_path": MAIN_ZIP_FILE},
        poke_interval=30, # 30s
        timeout=300,
        mode="poke"
    )

    wait_for_additional_file = PythonSensor(
        task_id="wait_for_additional_file",
        python_callable=check_file_exists,
        op_kwargs={"file_path": ADDITIONAL_ZIP_FILE},
        poke_interval=30,
        timeout=300,
        mode="poke"
    )

    # Task Staging
    stage_main_table = PythonOperator(
        task_id="stage_main_table",
        python_callable=extract_and_load_csv,
        op_kwargs={
            "zip_path": MAIN_ZIP_FILE,
            "csv_filename": "bank-full.csv",
            "table_name": "main",
            "postgres_conn_id": "postgres_default"
        }
    )

    # metadata
    stage_metadata_table = PythonOperator(
        task_id="stage_metadata_table",
        python_callable=extract_and_load_csv,
        op_kwargs={
            "zip_path": ADDITIONAL_ZIP_FILE,
            "csv_filename": "bank-additional-full.csv",
            "table_name": "metadata",
            "postgres_conn_id": "postgres_default"
        }
    )
    
    dbt_task_group = DbtTaskGroup(
        group_id="dbt_transformation",
        project_config=ProjectConfig(dbt_project_path=DBT_PROJECT_PATH),
        profile_config=profile_config,
    )

    make_excel = PythonOperator(
        task_id="make_excel_report",
        python_callable=generate_excel_report,
        op_kwargs={"postgres_conn_id": "postgres_default"}
    )

    # Integrated ML and PDF generation
    run_ml_and_pdf = PythonOperator(
        task_id="run_ml_and_generate_pdf",
        python_callable=run_ml_pipeline_with_report,
        op_kwargs={"postgres_conn_id": "postgres_default"}
    )

    # Alur
    [wait_for_main_file, wait_for_additional_file] >> stage_main_table
    [wait_for_main_file, wait_for_additional_file] >> stage_metadata_table
    [stage_main_table, stage_metadata_table] >> dbt_task_group
    dbt_task_group >> [run_ml_and_pdf, make_excel]